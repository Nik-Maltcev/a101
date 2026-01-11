"""Integration tests for API endpoints.

Tests the full API cycle: upload -> status -> download.
Uses mocked Redis and Celery for isolated testing.
"""

import io
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock
from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app
from app.models.schemas import JobStatus


@pytest.fixture
def client():
    """Create test client."""
    return TestClient(app)


@pytest.fixture
def sample_xlsx_bytes():
    """Create a sample xlsx file in memory."""
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws["A1"] = "ID"
    ws["B1"] = "Суть обращения"
    ws["C1"] = "Status"
    
    # Data rows
    ws["A2"] = "1"
    ws["B2"] = "Дефект 1, дефект 2"
    ws["C2"] = "Open"
    
    ws["A3"] = "2"
    ws["B3"] = "Дефект 3"
    ws["C3"] = "Closed"
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()
    
    return buffer.getvalue()


@pytest.fixture
def mock_redis():
    """Mock Redis client."""
    mock_client = MagicMock()
    mock_client.hgetall.return_value = {}
    mock_client.hset.return_value = True
    mock_client.expire.return_value = True
    return mock_client


class TestHealthEndpoint:
    """Tests for health check endpoint."""
    
    def test_health_check(self, client):
        """Test health endpoint returns healthy status."""
        response = client.get("/health")
        assert response.status_code == 200
        assert response.json() == {"status": "healthy"}


class TestRootEndpoint:
    """Tests for root endpoint."""
    
    def test_root_returns_web_ui(self, client):
        """Test root endpoint returns the web UI (HTML page)."""
        response = client.get("/")
        assert response.status_code == 200
        assert "text/html" in response.headers["content-type"]
        # Verify it contains expected HTML elements for the upload form
        assert b"<!DOCTYPE html>" in response.content or b"<html" in response.content


class TestCreateJobEndpoint:
    """Tests for POST /jobs endpoint."""
    
    @patch("app.api.jobs.get_redis_client")
    @patch("app.api.jobs.process_job")
    def test_create_job_success(
        self, mock_process_job, mock_get_redis, client, sample_xlsx_bytes, mock_redis
    ):
        """Test successful job creation with valid xlsx file."""
        mock_get_redis.return_value = mock_redis
        mock_process_job.delay.return_value = MagicMock()
        
        response = client.post(
            "/jobs",
            files={"file": ("test.xlsx", sample_xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "job_id" in data
        assert len(data["job_id"]) == 36  # UUID format
    
    def test_create_job_invalid_format(self, client):
        """Test job creation fails with non-xlsx file."""
        response = client.post(
            "/jobs",
            files={"file": ("test.txt", b"not an excel file", "text/plain")}
        )
        
        assert response.status_code == 400
        assert "xlsx" in response.json()["detail"].lower()
    
    def test_create_job_no_file(self, client):
        """Test job creation fails without file."""
        response = client.post("/jobs")
        
        assert response.status_code == 422  # Validation error


class TestGetJobStatusEndpoint:
    """Tests for GET /jobs/{job_id} endpoint."""
    
    @patch("app.api.jobs.get_redis_client")
    def test_get_job_status_not_found(self, mock_get_redis, client, mock_redis):
        """Test status returns 404 for non-existent job."""
        mock_redis.hgetall.return_value = {}
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/non-existent-job-id")
        
        assert response.status_code == 404
        assert "not found" in response.json()["detail"].lower()
    
    @patch("app.api.jobs.get_redis_client")
    def test_get_job_status_pending(self, mock_get_redis, client, mock_redis):
        """Test status returns pending job info."""
        mock_redis.hgetall.return_value = {
            "id": "test-job-id",
            "status": "pending",
            "progress": "0",
            "input_file": "/path/to/file.xlsx",
            "output_file": "",
            "error": "",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/test-job-id")
        
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "pending"
        assert data["progress"] == 0
        assert data["download_url"] is None
    
    @patch("app.api.jobs.get_redis_client")
    def test_get_job_status_completed(self, mock_get_redis, client, mock_redis, tmp_path):
        """Test status returns completed job with download URL."""
        output_file = tmp_path / "result.xlsx"
        output_file.touch()
        
        mock_redis.hgetall.return_value = {
            "id": "test-job-id",
            "status": "completed",
            "progress": "100",
            "input_file": "/path/to/file.xlsx",
            "output_file": str(output_file),
            "error": "",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/test-job-id")
        
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "completed"
        assert data["progress"] == 100
        assert data["download_url"] == "/jobs/test-job-id/download"


class TestDownloadEndpoint:
    """Tests for GET /jobs/{job_id}/download endpoint."""
    
    @patch("app.api.jobs.get_redis_client")
    def test_download_not_found(self, mock_get_redis, client, mock_redis):
        """Test download returns 404 for non-existent job."""
        mock_redis.hgetall.return_value = {}
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/non-existent-job-id/download")
        
        assert response.status_code == 404
    
    @patch("app.api.jobs.get_redis_client")
    def test_download_not_completed(self, mock_get_redis, client, mock_redis):
        """Test download returns 400 for non-completed job."""
        mock_redis.hgetall.return_value = {
            "id": "test-job-id",
            "status": "pending",
            "progress": "50",
            "input_file": "/path/to/file.xlsx",
            "output_file": "",
            "error": "",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/test-job-id/download")
        
        assert response.status_code == 400
        assert "not completed" in response.json()["detail"].lower()
    
    @patch("app.api.jobs.get_redis_client")
    def test_download_success(self, mock_get_redis, client, mock_redis, tmp_path, sample_xlsx_bytes):
        """Test successful file download."""
        # Create actual output file
        output_file = tmp_path / "result.xlsx"
        output_file.write_bytes(sample_xlsx_bytes)
        
        mock_redis.hgetall.return_value = {
            "id": "test-job-id",
            "status": "completed",
            "progress": "100",
            "input_file": "/path/to/file.xlsx",
            "output_file": str(output_file),
            "error": "",
            "created_at": datetime.utcnow().isoformat(),
            "updated_at": datetime.utcnow().isoformat(),
        }
        mock_get_redis.return_value = mock_redis
        
        response = client.get("/jobs/test-job-id/download")
        
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
