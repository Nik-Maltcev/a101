"""Tests for ExcelReader service."""

import pytest
from pathlib import Path
from openpyxl import Workbook
import tempfile
import os

from app.services.excel_reader import (
    ExcelReader,
    ExcelReaderError,
    CommentColumnNotFoundError,
)


@pytest.fixture
def excel_reader():
    """Create ExcelReader instance."""
    return ExcelReader()


@pytest.fixture
def sample_xlsx_with_comments(tmp_path):
    """Create a sample xlsx file with КОММЕНТАРИЙ column."""
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws["A1"] = "ID"
    ws["B1"] = "КОММЕНТАРИЙ"
    ws["C1"] = "Status"
    
    # Data rows
    ws["A2"] = "1"
    ws["B2"] = "Дефект 1, дефект 2"
    ws["C2"] = "Open"
    
    ws["A3"] = "2"
    ws["B3"] = "Дефект 3"
    ws["C3"] = "Closed"
    
    ws["A4"] = "3"
    ws["B4"] = "нет замечаний"
    ws["C4"] = "Done"
    
    file_path = tmp_path / "test_comments.xlsx"
    wb.save(file_path)
    wb.close()
    
    return file_path


@pytest.fixture
def sample_xlsx_without_comments(tmp_path):
    """Create a sample xlsx file without КОММЕНТАРИЙ column."""
    wb = Workbook()
    ws = wb.active
    
    ws["A1"] = "ID"
    ws["B1"] = "Description"
    ws["C1"] = "Status"
    
    ws["A2"] = "1"
    ws["B2"] = "Some text"
    ws["C2"] = "Open"
    
    file_path = tmp_path / "test_no_comments.xlsx"
    wb.save(file_path)
    wb.close()
    
    return file_path


class TestExcelReader:
    """Tests for ExcelReader class."""
    
    def test_read_file_success(self, excel_reader, sample_xlsx_with_comments):
        """Test reading a valid xlsx file with КОММЕНТАРИЙ column."""
        rows = excel_reader.read_file(sample_xlsx_with_comments)
        
        assert len(rows) == 3
        assert rows[0]["ID"] == "1"
        assert rows[0]["КОММЕНТАРИЙ"] == "Дефект 1, дефект 2"
        assert rows[0]["Status"] == "Open"
        
        assert rows[1]["ID"] == "2"
        assert rows[1]["КОММЕНТАРИЙ"] == "Дефект 3"
        
        assert rows[2]["КОММЕНТАРИЙ"] == "нет замечаний"
    
    def test_find_comment_column(self, excel_reader, sample_xlsx_with_comments):
        """Test finding КОММЕНТАРИЙ column index."""
        from openpyxl import load_workbook
        
        wb = load_workbook(sample_xlsx_with_comments, read_only=True)
        sheet = wb.active
        
        col_idx = excel_reader.find_comment_column(sheet)
        
        assert col_idx == 1  # 0-based index, B column
        wb.close()
    
    def test_find_comment_column_not_found(self, excel_reader, sample_xlsx_without_comments):
        """Test when КОММЕНТАРИЙ column is not found."""
        from openpyxl import load_workbook
        
        wb = load_workbook(sample_xlsx_without_comments, read_only=True)
        sheet = wb.active
        
        col_idx = excel_reader.find_comment_column(sheet)
        
        assert col_idx is None
        wb.close()
    
    def test_read_file_missing_comment_column(self, excel_reader, sample_xlsx_without_comments):
        """Test reading file without КОММЕНТАРИЙ column raises error."""
        with pytest.raises(CommentColumnNotFoundError):
            excel_reader.read_file(sample_xlsx_without_comments)
    
    def test_read_file_not_found(self, excel_reader):
        """Test reading non-existent file raises error."""
        with pytest.raises(ExcelReaderError, match="File not found"):
            excel_reader.read_file("/nonexistent/path/file.xlsx")
    
    def test_read_file_invalid_format(self, excel_reader, tmp_path):
        """Test reading non-xlsx file raises error."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("not an excel file")
        
        with pytest.raises(ExcelReaderError, match="Invalid file format"):
            excel_reader.read_file(txt_file)
    
    def test_read_file_empty_rows_skipped(self, excel_reader, tmp_path):
        """Test that empty rows are skipped."""
        wb = Workbook()
        ws = wb.active
        
        ws["A1"] = "ID"
        ws["B1"] = "КОММЕНТАРИЙ"
        
        ws["A2"] = "1"
        ws["B2"] = "Comment 1"
        
        # Row 3 is empty
        ws["A3"] = None
        ws["B3"] = None
        
        ws["A4"] = "2"
        ws["B4"] = "Comment 2"
        
        file_path = tmp_path / "test_empty_rows.xlsx"
        wb.save(file_path)
        wb.close()
        
        rows = excel_reader.read_file(file_path)
        
        assert len(rows) == 2
        assert rows[0]["ID"] == "1"
        assert rows[1]["ID"] == "2"
