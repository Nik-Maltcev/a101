"""API endpoints for Domyland integration."""

import uuid
from pathlib import Path
from typing import Optional
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel

from app.config import settings
from app.services.domyland_client import DomylandClient, DomylandAuthError, DomylandClientError
from app.services.domyland_export import DomylandExportService
from app.services.domyland_processor import process_domyland_data

router = APIRouter(prefix="/domyland", tags=["domyland"])


# Store tokens in memory (in production use Redis/DB)
_tokens: dict[str, dict] = {}


class AuthRequest(BaseModel):
    """Authentication request."""
    email: str
    password: str
    tenant_name: str = "a101"


class AuthResponse(BaseModel):
    """Authentication response."""
    success: bool
    message: str
    session_id: Optional[str] = None
    available_exports: Optional[list[dict]] = None  # List of available export types


class ExportRequest(BaseModel):
    """Export request."""
    session_id: str
    export_type: str  # buildings, customers, places, orders, payments
    # Optional filters
    building_id: Optional[int] = None
    created_at: Optional[str] = None  # DD.MM.YYYY-DD.MM.YYYY
    updated_at: Optional[str] = None
    service_ids: Optional[list[int]] = None  # Filter by service IDs (multiple)


class ExportResponse(BaseModel):
    """Export response."""
    success: bool
    message: str
    download_url: Optional[str] = None
    record_count: Optional[int] = None


@router.post("/auth", response_model=AuthResponse)
async def authenticate(request: AuthRequest) -> AuthResponse:
    """
    Authenticate with Domyland API.
    
    Returns a session_id to use for subsequent requests.
    Also checks which export types are available for this user.
    """
    client = DomylandClient()
    
    try:
        token = await client.authenticate(
            email=request.email,
            password=request.password,
            tenant_name=request.tenant_name,
        )
        
        # Check available permissions
        available_exports = await _check_permissions(client)
        
        # Generate session ID and store token
        session_id = str(uuid.uuid4())
        _tokens[session_id] = {
            "token": token,
            "tenant_name": request.tenant_name,
            "created_at": datetime.utcnow().isoformat(),
            "available_exports": [e["id"] for e in available_exports],
        }
        
        return AuthResponse(
            success=True,
            message=f"Авторизация успешна. Доступно {len(available_exports)} типов экспорта.",
            session_id=session_id,
            available_exports=available_exports,
        )
        
    except DomylandAuthError as e:
        return AuthResponse(
            success=False,
            message=str(e),
        )
    finally:
        await client.close()


async def _check_permissions(client: DomylandClient) -> list[dict]:
    """Check which API endpoints are accessible for the authenticated user."""
    # Only check main endpoints, skip slow/experimental ones
    all_types = [
        {"id": "buildings", "name": "Объекты/здания", "endpoint": "buildings"},
        {"id": "places", "name": "Помещения", "endpoint": "places"},
        {"id": "orders", "name": "Заявки", "endpoint": "orders"},
        {"id": "orders_raw", "name": "Заявки (все поля)", "endpoint": "orders"},
        {"id": "customers", "name": "Клиенты", "endpoint": "customers"},
        {"id": "payments", "name": "Платежи", "endpoint": "payments"},
    ]
    
    available = []
    for export_type in all_types:
        try:
            # Try to fetch first page to check access (with short timeout)
            await client._request("GET", export_type["endpoint"], params={"fromRow": 0})
            available.append({
                "id": export_type["id"],
                "name": export_type["name"],
            })
        except DomylandClientError:
            # No access to this endpoint
            pass
        except Exception:
            # Timeout or other error - skip
            pass
    
    return available


@router.get("/services/{session_id}")
async def get_services(session_id: str, from_row: int = 0):
    """Get list of available services for filtering (paginated)."""
    session = _tokens.get(session_id)
    if not session:
        raise HTTPException(status_code=401, detail="Сессия не найдена")
    
    client = DomylandClient()
    client.set_token(session["token"])
    
    try:
        services, next_row = await client.get_services(from_row)
        # Return simplified list with id and title
        return {
            "services": [
                {"id": s.get("id"), "title": s.get("title") or s.get("internalTitle") or f"Service {s.get('id')}"}
                for s in services
            ],
            "next_row": next_row,
            "has_more": next_row != -1
        }
    except DomylandAuthError:
        _tokens.pop(session_id, None)
        raise HTTPException(status_code=401, detail="Сессия истекла")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        await client.close()


@router.post("/export", response_model=ExportResponse)
async def export_data(request: ExportRequest) -> ExportResponse:
    """
    Export data from Domyland to Excel file.
    
    Supported export_type values:
    - buildings: Объекты/здания
    - customers: Клиенты
    - places: Помещения
    - orders: Заявки со счетами
    - payments: Платежи
    """
    # Check session
    session = _tokens.get(request.session_id)
    if not session:
        return ExportResponse(
            success=False,
            message="Сессия не найдена. Авторизуйтесь заново.",
        )
    
    # Create client with stored token
    client = DomylandClient()
    client.set_token(session["token"])
    export_service = DomylandExportService(client)
    
    # Generate output filename
    file_id = str(uuid.uuid4())
    output_path = settings.RESULTS_DIR / f"domyland_{request.export_type}_{file_id}.xlsx"
    
    try:
        if request.export_type == "buildings":
            await export_service.export_buildings(output_path, request.updated_at)
        elif request.export_type == "customers":
            await export_service.export_customers(output_path, request.updated_at)
        elif request.export_type == "places":
            await export_service.export_places(output_path, request.updated_at)
        elif request.export_type == "orders":
            await export_service.export_orders(
                output_path,
                building_id=request.building_id,
                created_at=request.created_at,
                service_ids=request.service_ids,
            )
        elif request.export_type == "payments":
            await export_service.export_payments(output_path, request.created_at)
        elif request.export_type == "orders_raw":
            await export_service.export_orders_raw(
                output_path,
                building_id=request.building_id,
                created_at=request.created_at,
            )
        elif request.export_type == "acceptance_results":
            await export_service.export_acceptance_results(
                output_path,
                building_id=request.building_id,
            )
        elif request.export_type == "acceptance_defects":
            await export_service.export_acceptance_defects(
                output_path,
                building_id=request.building_id,
            )
        elif request.export_type == "export_columns":
            await export_service.export_columns_list(output_path)
        else:
            return ExportResponse(
                success=False,
                message=f"Неизвестный тип экспорта: {request.export_type}",
            )
        
        # Count records in file
        from openpyxl import load_workbook
        wb = load_workbook(output_path, read_only=True)
        record_count = wb.active.max_row - 1  # Minus header
        wb.close()
        
        return ExportResponse(
            success=True,
            message=f"Экспорт завершён: {record_count} записей",
            download_url=f"/domyland/download/{file_id}?type={request.export_type}",
            record_count=record_count,
        )
        
    except DomylandAuthError:
        # Token expired, remove session
        _tokens.pop(request.session_id, None)
        return ExportResponse(
            success=False,
            message="Сессия истекла. Авторизуйтесь заново.",
        )
    except DomylandClientError as e:
        return ExportResponse(
            success=False,
            message=f"Ошибка API: {e}",
        )
    except Exception as e:
        return ExportResponse(
            success=False,
            message=f"Ошибка экспорта: {e}",
        )
    finally:
        await client.close()


@router.get("/download/{file_id}")
async def download_export(
    file_id: str,
    type: str = Query(..., description="Export type"),
) -> FileResponse:
    """Download exported Excel file."""
    output_path = settings.RESULTS_DIR / f"domyland_{type}_{file_id}.xlsx"
    
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Файл не найден")
    
    return FileResponse(
        path=str(output_path),
        filename=f"domyland_{type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/export-types")
async def get_export_types():
    """Get available export types."""
    return {
        "types": [
            {"id": "buildings", "name": "Объекты/здания", "description": "Список зданий и объектов"},
            {"id": "customers", "name": "Клиенты", "description": "Список клиентов"},
            {"id": "places", "name": "Помещения", "description": "Список помещений/квартир"},
            {"id": "orders", "name": "Заявки со счетами", "description": "Заявки с информацией об оплате"},
            {"id": "payments", "name": "Платежи", "description": "История платежей"},
        ]
    }



class ProcessRequest(BaseModel):
    """Request to process Domyland export file."""
    session_id: str
    file_id: str
    export_type: str = "orders"


class ProcessResponse(BaseModel):
    """Response from processing."""
    success: bool
    message: str
    download_url: Optional[str] = None
    original_count: Optional[int] = None
    expanded_count: Optional[int] = None


@router.post("/process", response_model=ProcessResponse)
async def process_domyland_file(request: ProcessRequest) -> ProcessResponse:
    """
    Process exported Domyland file - split defects by "|" delimiter.
    
    This is a separate processor that:
    1. Uses "|" as delimiter (not LLM)
    2. Filters out location-only values (Окно, Стена, etc.)
    3. Expands rows - one row per defect
    """
    # Check session
    session = _tokens.get(request.session_id)
    if not session:
        return ProcessResponse(
            success=False,
            message="Сессия не найдена. Авторизуйтесь заново.",
        )
    
    # Find source file
    source_path = settings.RESULTS_DIR / f"domyland_{request.export_type}_{request.file_id}.xlsx"
    if not source_path.exists():
        return ProcessResponse(
            success=False,
            message="Исходный файл не найден. Сначала выполните экспорт.",
        )
    
    try:
        from openpyxl import load_workbook, Workbook
        
        # Read source file
        wb = load_workbook(source_path, read_only=True)
        ws = wb.active
        
        # Get headers and data
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ProcessResponse(success=False, message="Файл пустой")
        
        headers = list(rows[0])
        data = []
        for row in rows[1:]:
            data.append(dict(zip(headers, row)))
        
        wb.close()
        
        original_count = len(data)
        
        # Process data - split defects
        processed_data = process_domyland_data(data)
        expanded_count = len(processed_data)
        
        # Write processed file
        processed_file_id = str(uuid.uuid4())
        output_path = settings.RESULTS_DIR / f"domyland_processed_{processed_file_id}.xlsx"
        
        # Add "Дефект" column to headers if not present
        output_headers = headers.copy()
        if "Дефект" not in output_headers:
            output_headers.append("Дефект")
        
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Processed"
        
        # Write headers
        for col, header in enumerate(output_headers, 1):
            ws_out.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, record in enumerate(processed_data, 2):
            for col_idx, header in enumerate(output_headers, 1):
                value = record.get(header, "")
                ws_out.cell(row=row_idx, column=col_idx, value=value)
        
        wb_out.save(output_path)
        
        return ProcessResponse(
            success=True,
            message=f"Обработано: {original_count} → {expanded_count} строк",
            download_url=f"/domyland/download-processed/{processed_file_id}",
            original_count=original_count,
            expanded_count=expanded_count,
        )
        
    except Exception as e:
        return ProcessResponse(
            success=False,
            message=f"Ошибка обработки: {e}",
        )


@router.get("/download-processed/{file_id}")
async def download_processed(file_id: str) -> FileResponse:
    """Download processed Domyland file."""
    output_path = settings.RESULTS_DIR / f"domyland_processed_{file_id}.xlsx"
    
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Файл не найден")
    
    return FileResponse(
        path=str(output_path),
        filename=f"domyland_processed_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
