"""API endpoints for job management."""

import asyncio
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

from fastapi import APIRouter, BackgroundTasks, File, HTTPException, UploadFile
from fastapi.responses import FileResponse

from app.config import settings
from app.models import Job, JobResponse, JobStatus, JobStatusResponse

router = APIRouter(prefix="/jobs", tags=["jobs"])

# In-memory job storage (for simple deployment without Redis)
_jobs_storage: Dict[str, dict] = {}


def get_job(job_id: str) -> Job | None:
    """Get job from storage."""
    job_data = _jobs_storage.get(job_id)
    if not job_data:
        return None
    return Job(
        id=job_data["id"],
        status=JobStatus(job_data["status"]),
        progress=int(job_data["progress"]),
        input_file=job_data["input_file"],
        output_file=job_data.get("output_file") or None,
        error=job_data.get("error") or None,
        created_at=datetime.fromisoformat(job_data["created_at"]),
        updated_at=datetime.fromisoformat(job_data["updated_at"]),
    )


def save_job(job: Job) -> None:
    """Save job to storage."""
    job_data = {
        "id": job.id,
        "status": job.status.value,
        "progress": str(job.progress),
        "input_file": job.input_file,
        "output_file": job.output_file or "",
        "error": job.error or "",
        "created_at": job.created_at.isoformat(),
        "updated_at": job.updated_at.isoformat(),
    }
    _jobs_storage[job.id] = job_data


def update_job_status(
    job_id: str,
    status: JobStatus,
    progress: int = 0,
    output_file: Optional[str] = None,
    error: Optional[str] = None,
) -> None:
    """Update job status in storage."""
    if job_id not in _jobs_storage:
        return
    
    _jobs_storage[job_id]["status"] = status.value
    _jobs_storage[job_id]["progress"] = str(progress)
    _jobs_storage[job_id]["updated_at"] = datetime.utcnow().isoformat()
    
    if output_file is not None:
        _jobs_storage[job_id]["output_file"] = output_file
    if error is not None:
        _jobs_storage[job_id]["error"] = error


async def process_job_async(job_id: str, file_path: str) -> None:
    """
    Process job asynchronously.
    
    Steps:
    1. Read xlsx file via ExcelReader
    2. Split comments into defects via SplitService (LLM)
    3. Expand rows (one row per defect)
    4. Classify defects via ClassifyService (LLM)
    5. Write result file via ExcelWriter
    """
    import logging
    from app.services.excel_reader import ExcelReader, ExcelReaderError
    from app.services.split_service import SplitService
    from app.services.expand_service import expand_rows
    from app.services.classify_service import ClassifyService
    from app.services.excel_writer import ExcelWriter, get_output_path
    from app.services.category_index import CategoryIndex
    from app.services.llm_client import LLMClient
    
    logger = logging.getLogger(__name__)
    logger.info(f"Job {job_id}: Starting async processing")
    llm_client = None
    
    try:
        # Initialize services
        llm_client = LLMClient()
        category_index = CategoryIndex(settings.CATEGORIES_FILE)
        category_index.build_index()
        
        split_service = SplitService(llm_client)
        classify_service = ClassifyService(llm_client, category_index)
        excel_reader = ExcelReader()
        excel_writer = ExcelWriter()
        
        # Step 1: Read xlsx file
        logger.info(f"Job {job_id}: Reading file {file_path}")
        update_job_status(job_id, JobStatus.PENDING, progress=5)
        
        rows = excel_reader.read_file(file_path)
        total_rows = len(rows)
        logger.info(f"Job {job_id}: Read {total_rows} rows")
        
        if total_rows == 0:
            output_path = get_output_path(job_id, settings.RESULTS_DIR)
            excel_writer.write_result([], str(output_path))
            update_job_status(job_id, JobStatus.COMPLETED, progress=100, output_file=str(output_path))
            return
        
        # Extract comments by concatenating valueString + valueText
        def get_comment(row: dict) -> str:
            value_string = ""
            value_text = ""
            
            # Find valueString and valueText (case-insensitive)
            for key in row.keys():
                if key.upper() == "VALUESTRING":
                    value_string = row.get(key, "") or ""
                elif key.upper() == "VALUETEXT":
                    value_text = row.get(key, "") or ""
            
            # Concatenate with space if both exist
            if value_string and value_text:
                return f"{value_string} {value_text}"
            elif value_string:
                return value_string
            elif value_text:
                return value_text
            else:
                return ""
        
        comments = [get_comment(row) for row in rows]
        
        # Log first few comments for debugging
        logger.info(f"Job {job_id}: First 3 comments to split:")
        for i, comment in enumerate(comments[:3]):
            logger.info(f"  Comment {i+1}: {comment[:200]}...")
        
        # Step 2: Split comments
        logger.info(f"Job {job_id}: Splitting {len(comments)} comments")
        update_job_status(job_id, JobStatus.SPLITTING, progress=10)
        
        defects_per_row = await split_service.split_batch(comments)
        total_defects = sum(len(d) for d in defects_per_row)
        logger.info(f"Job {job_id}: Found {total_defects} defects")
        
        update_job_status(job_id, JobStatus.SPLITTING, progress=40)
        
        # Step 3: Expand rows
        logger.info(f"Job {job_id}: Expanding rows")
        expanded_rows = expand_rows(rows, defects_per_row)
        logger.info(f"Job {job_id}: Expanded to {len(expanded_rows)} rows")
        
        update_job_status(job_id, JobStatus.CLASSIFYING, progress=50)
        
        # Step 4: Classify defects
        if expanded_rows:
            logger.info(f"Job {job_id}: Classifying {len(expanded_rows)} defects")
            defect_texts = [row.defect_text for row in expanded_rows]
            categories = await classify_service.classify_batch(defect_texts)
            
            for i, category in enumerate(categories):
                expanded_rows[i].category = category
            
            logger.info(f"Job {job_id}: Classification complete")
        
        update_job_status(job_id, JobStatus.CLASSIFYING, progress=90)
        
        # Step 5: Write result
        logger.info(f"Job {job_id}: Writing result file")
        output_path = get_output_path(job_id, settings.RESULTS_DIR)
        original_headers = list(rows[0].keys()) if rows else None
        
        excel_writer.write_result(expanded_rows, str(output_path), original_headers=original_headers)
        
        logger.info(f"Job {job_id}: Result saved to {output_path}")
        update_job_status(job_id, JobStatus.COMPLETED, progress=100, output_file=str(output_path))
        
    except ExcelReaderError as e:
        logger.error(f"Job {job_id}: Excel read error - {e}")
        update_job_status(job_id, JobStatus.FAILED, error=f"Failed to read file: {e}")
    except Exception as e:
        logger.error(f"Job {job_id}: Processing error - {e}", exc_info=True)
        update_job_status(job_id, JobStatus.FAILED, error=f"Processing failed: {e}")
    finally:
        if llm_client is not None:
            await llm_client.close()


def run_process_job(job_id: str, file_path: str):
    """Run async job processing in background."""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Starting background job processing for {job_id}")
    
    try:
        asyncio.run(process_job_async(job_id, file_path))
        logger.info(f"Background job {job_id} completed")
    except Exception as e:
        logger.error(f"Background job {job_id} failed with error: {e}", exc_info=True)
        update_job_status(job_id, JobStatus.FAILED, error=f"Background processing failed: {e}")


@router.post("", response_model=JobResponse)
async def create_job(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
) -> JobResponse:
    """
    Upload an Excel file for processing.
    
    Accepts only .xlsx files. Creates a job and starts processing in background.
    """
    # Validate file format
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in settings.ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")
    
    # Check file size
    content = await file.read()
    if len(content) > settings.MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File size exceeds limit of {settings.MAX_FILE_SIZE // (1024*1024)} MB"
        )
    
    # Generate unique job_id
    job_id = str(uuid.uuid4())
    
    # Save file
    file_path = settings.UPLOADS_DIR / f"{job_id}.xlsx"
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Create job record
    now = datetime.utcnow()
    job = Job(
        id=job_id,
        status=JobStatus.PENDING,
        progress=0,
        input_file=str(file_path),
        output_file=None,
        error=None,
        created_at=now,
        updated_at=now,
    )
    save_job(job)
    
    # Start processing in background
    background_tasks.add_task(run_process_job, job_id, str(file_path))
    
    return JobResponse(job_id=job_id)


@router.get("/{job_id}", response_model=JobStatusResponse)
async def get_job_status(job_id: str) -> JobStatusResponse:
    """
    Get the status of a processing job.
    
    Returns status, progress percentage, and download URL when completed.
    
    Requirements: 8.1, 8.2
    """
    job = get_job(job_id)
    
    # Job not found (Requirement 8.4)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Build download URL if completed (Requirement 8.2)
    download_url = None
    if job.status == JobStatus.COMPLETED and job.output_file:
        download_url = f"/jobs/{job_id}/download"
    
    return JobStatusResponse(
        status=job.status,
        progress=job.progress,
        download_url=download_url,
        error=job.error,
    )


@router.get("/{job_id}/download")
async def download_result(job_id: str) -> FileResponse:
    """
    Download the processed result file.
    
    Only available when job status is 'completed'.
    
    Requirements: 8.3, 8.4
    """
    job = get_job(job_id)
    
    # Job not found
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Job not completed yet
    if job.status != JobStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="Job is not completed yet")
    
    # Check output file exists
    if not job.output_file:
        raise HTTPException(status_code=404, detail="Result file not found")
    
    output_path = Path(job.output_file)
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Result file not found")
    
    # Return file (Requirement 8.3)
    return FileResponse(
        path=str(output_path),
        filename=f"{job_id}_processed.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/{job_id}/analytics")
async def get_analytics(job_id: str):
    """
    Get analytics data for a completed job.
    
    Returns category distribution, column info, and raw data.
    """
    from openpyxl import load_workbook
    from collections import Counter
    import logging
    
    logger = logging.getLogger(__name__)
    
    job = get_job(job_id)
    
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.status != JobStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="Job is not completed yet")
    
    if not job.output_file:
        raise HTTPException(status_code=404, detail="Result file not found")
    
    output_path = Path(job.output_file)
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Result file not found")
    
    # Read the processed file
    workbook = load_workbook(filename=output_path, read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    logger.info(f"Analytics: Found headers: {headers}")
    
    # Find category column FIRST (case-insensitive)
    category_column = None
    for h in headers:
        if "категория" in h.lower():
            category_column = h
            break
    
    logger.info(f"Analytics: Category column = {category_column}")
    
    # Read all data
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                # Convert to string, handle None
                row_data[headers[i]] = str(value).strip() if value is not None else ""
        if any(v for v in row_data.values() if v):  # Skip empty rows
            rows.append(row_data)
    
    workbook.close()
    
    logger.info(f"Analytics: Read {len(rows)} rows")
    
    # Calculate category distribution
    category_distribution = []
    total_categories = 0
    
    if category_column and rows:
        # Get all non-empty category values
        categories = []
        # Log first few rows for debugging
        for i, row in enumerate(rows[:5]):
            cat_value = row.get(category_column, "")
            logger.info(f"Analytics: Row {i} category raw value: '{cat_value}' (type: {type(cat_value).__name__})")
        
        for row in rows:
            cat_value = row.get(category_column, "")
            if cat_value and cat_value != "None" and cat_value.strip():
                categories.append(cat_value)
        
        logger.info(f"Analytics: Found {len(categories)} non-empty category values out of {len(rows)} rows")
        
        if categories:
            category_counts = Counter(categories)
            total_categories = len(category_counts)
            total = len(categories)
            category_distribution = [
                {"category": cat, "count": count, "percentage": round(count / total * 100, 1)}
                for cat, count in category_counts.most_common(20)
            ]
        
        logger.info(f"Analytics: Found {total_categories} unique categories")
    
    # Get unique values for each column (for filters)
    column_values = {}
    for header in headers:
        values = set()
        for row in rows:
            val = row.get(header, "")
            if val and val != "None":
                values.add(val)
        column_values[header] = sorted(list(values))[:50]
    
    return {
        "total_rows": len(rows),
        "total_categories": total_categories,
        "headers": headers,
        "category_column": category_column,
        "category_distribution": category_distribution,
        "column_values": column_values,
        "data": rows[:500],
    }
