"""Service for exporting Domyland data to Excel."""

import logging
import re
from pathlib import Path
from typing import Optional
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook

from app.services.domyland_client import DomylandClient

logger = logging.getLogger(__name__)

# Regex to find image URLs from domyland uploads
# Matches both full URLs with extension and truncated URLs without extension
PHOTO_URL_PATTERN = re.compile(
    r'https?://uploads\.domyland\.com/[a-zA-Z0-9_-]+(?:\.(jpeg|jpg|png|gif))?',
    re.IGNORECASE
)

# Moscow timezone (UTC+3)
MOSCOW_TZ = timezone(timedelta(hours=3))


class DomylandExportService:
    """Service for exporting data from Domyland API to Excel files."""
    
    def __init__(self, client: DomylandClient):
        self.client = client
    
    def _flatten_dict(self, d: dict, parent_key: str = '', sep: str = '_') -> dict:
        """Flatten nested dictionary."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep).items())
            elif isinstance(v, list):
                # Convert list to string representation
                if v and isinstance(v[0], dict):
                    # For list of dicts, take first item's relevant field
                    items.append((new_key, str(v)))
                else:
                    items.append((new_key, ", ".join(str(x) for x in v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def _write_to_excel(self, data: list[dict], output_path: Path, sheet_name: str = "Data") -> Path:
        """Write list of dicts to Excel file."""
        if not data:
            logger.warning("No data to export")
            # Create empty file with headers
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            wb.save(output_path)
            return output_path
        
        # Flatten all records
        flat_data = [self._flatten_dict(record) for record in data]
        
        # Get all unique headers
        all_headers = set()
        for record in flat_data:
            all_headers.update(record.keys())
        headers = sorted(list(all_headers))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, record in enumerate(flat_data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, "")
                # Convert timestamps to readable dates
                if isinstance(value, int) and value > 1000000000 and value < 2000000000:
                    try:
                        # Convert to Moscow time
                        value = datetime.fromtimestamp(value, tz=MOSCOW_TZ).strftime("%d.%m.%Y %H:%M")
                    except:
                        pass
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
        logger.info(f"Exported {len(data)} records to {output_path}")
        return output_path
    
    async def export_buildings(self, output_path: Path, updated_at: Optional[str] = None) -> Path:
        """Export buildings to Excel."""
        data = await self.client.get_buildings(updated_at)
        return self._write_to_excel(data, output_path, "Buildings")
    
    async def export_customers(self, output_path: Path, updated_at: Optional[str] = None) -> Path:
        """Export customers to Excel."""
        data = await self.client.get_customers(updated_at)
        return self._write_to_excel(data, output_path, "Customers")
    
    async def export_places(self, output_path: Path, updated_at: Optional[str] = None) -> Path:
        """Export places to Excel."""
        data = await self.client.get_places(updated_at)
        return self._write_to_excel(data, output_path, "Places")
    
    async def export_orders(
        self,
        output_path: Path,
        building_id: Optional[int] = None,
        created_at: Optional[str] = None,
        service_ids: Optional[list[int]] = None,
        fetch_full_details: bool = True,
    ) -> Path:
        """Export orders to Excel.
        
        Exports fields: id, address, title, valueString, valueText, extId, createdAt
        
        Field mapping:
        - valueString: answers from orderElements (dropdown/checkbox values)
        - valueText: customerSummary (contains detailed defect descriptions)
        
        Args:
            service_ids: Filter by service IDs (multiple). If provided, only orders
                        with matching serviceId will be exported.
            fetch_full_details: If True, fetch full details for each order to get
                               complete customerSummary (not truncated). Default True.
        """
        raw_data = await self.client.get_orders_with_invoices(
            building_id=building_id,
            created_at=created_at,
        )
        
        logger.info(f"Got {len(raw_data)} orders from API")
        
        # Log actual serviceIds in data BEFORE filtering
        if raw_data:
            actual_service_ids = set(order.get("serviceId") for order in raw_data)
            logger.info(f"Unique serviceIds in ALL orders: {len(actual_service_ids)} different IDs")
            logger.info(f"First 20 actual serviceIds: {list(actual_service_ids)[:20]}")
        
        # Filter by service_ids if provided
        if service_ids:
            before_filter = len(raw_data)
            service_ids_set = set(service_ids)
            logger.info(f"Requested service_ids: {list(service_ids_set)[:10]}... (total {len(service_ids_set)})")
            
            # Check intersection
            if raw_data:
                actual_service_ids = set(order.get("serviceId") for order in raw_data)
                intersection = actual_service_ids & service_ids_set
                logger.info(f"Intersection (matching IDs): {intersection}")
            
            raw_data = [order for order in raw_data if order.get("serviceId") in service_ids_set]
            logger.info(f"Filtered by service_ids: {before_filter} -> {len(raw_data)} orders")
        
        # Fetch full details for each order if requested
        if fetch_full_details and raw_data:
            logger.info(f"Fetching full details for {len(raw_data)} orders...")
            import asyncio
            
            async def fetch_order_details(order: dict) -> dict:
                """Fetch full details and merge with order."""
                order_id = order.get("id")
                if not order_id:
                    return order
                
                details = await self.client.get_order_details(order_id)
                if details:
                    # Log what we got from list vs details endpoint
                    old_summary = order.get("customerSummary") or ""
                    new_summary = details.get("customerSummary") or ""
                    old_len = len(old_summary)
                    new_len = len(new_summary)
                    
                    # Always log for debugging
                    logger.info(f"Order {order_id}: LIST customerSummary ({old_len} chars): {old_summary[:200]}...")
                    logger.info(f"Order {order_id}: DETAILS customerSummary ({new_len} chars): {new_summary[:200]}...")
                    
                    if new_len > old_len:
                        logger.info(f"Order {order_id}: customerSummary EXTENDED {old_len} -> {new_len} chars")
                    elif new_len < old_len:
                        logger.warning(f"Order {order_id}: customerSummary SHORTER in details! {old_len} -> {new_len}")
                    elif new_len == old_len and old_len > 0:
                        logger.info(f"Order {order_id}: customerSummary SAME length ({old_len} chars)")
                    
                    # Merge details into order, preferring details for customerSummary
                    merged = {**order, **details}
                    return merged
                return order
            
            # Process in batches to avoid overwhelming the API
            batch_size = 10
            enriched_data = []
            for i in range(0, len(raw_data), batch_size):
                batch = raw_data[i:i + batch_size]
                tasks = [fetch_order_details(order) for order in batch]
                results = await asyncio.gather(*tasks)
                enriched_data.extend(results)
                logger.info(f"Fetched details for orders {i+1}-{min(i+batch_size, len(raw_data))}/{len(raw_data)}")
            
            raw_data = enriched_data
        
        # Log first order structure for debugging
        if raw_data:
            first_order = raw_data[0]
            logger.info(f"First order keys: {list(first_order.keys())}")
            logger.info(f"First order id: {first_order.get('id')}")
            customer_summary = first_order.get('customerSummary', '')
            logger.info(f"First order customerSummary length: {len(customer_summary) if customer_summary else 0}")
            logger.info(f"First order customerSummary: {customer_summary[:500] if customer_summary else 'EMPTY'}...")
            logger.info(f"First order orderElements: {first_order.get('orderElements', [])}")
        
        # Transform data to extract only needed fields
        data = []
        for order in raw_data:
            order_elements = order.get("orderElements", [])
            
            # Collect all values from orderElements
            value_strings = []
            for elem in order_elements:
                val_title = elem.get("valueTitle")
                if val_title:
                    val_str = str(val_title).strip()
                    if val_str:
                        value_strings.append(val_str)
            
            # Remove duplicates while preserving order
            value_strings = list(dict.fromkeys(value_strings))
            
            # customerSummary contains the detailed defect text!
            raw_value_text = order.get("customerSummary") or ""
            
            # Log full length for debugging truncation issues
            order_id = order.get("id")
            if len(raw_value_text) > 500:
                logger.info(f"Order {order_id}: customerSummary length = {len(raw_value_text)} chars")
                logger.info(f"Order {order_id}: customerSummary FULL TEXT: {raw_value_text}")
            
            # Extract photo URLs and remove them from valueText
            # Use finditer to get full match objects
            full_photo_urls = [m.group(0) for m in PHOTO_URL_PATTERN.finditer(raw_value_text)]
            
            # Remove photo URLs from text
            value_text = PHOTO_URL_PATTERN.sub('', raw_value_text).strip()
            # Clean up extra spaces/newlines left after URL removal
            value_text = re.sub(r'\s+', ' ', value_text).strip()
            
            # Join photo URLs with newline for the Фото column
            photos = '\n'.join(full_photo_urls) if full_photo_urls else ""
            
            # Build address
            address = order.get("placeAddress") or order.get("buildingTitle") or ""
            
            # Convert timestamp
            created_at_ts = order.get("createdAt")
            created_at_str = ""
            if created_at_ts and isinstance(created_at_ts, int):
                try:
                    # Convert to Moscow time
                    created_at_str = datetime.fromtimestamp(created_at_ts, tz=MOSCOW_TZ).strftime("%d.%m.%Y %H:%M")
                except:
                    created_at_str = str(created_at_ts)
            
            # Format orderElements - extract only comments (fields containing "комментарий")
            order_elements_text = ""
            if order_elements:
                comments_parts = []
                for elem in order_elements:
                    q = elem.get("elementTitle", "")
                    a = elem.get("valueTitle", "")
                    # Only include comment fields
                    if q and a and "комментарий" in q.lower():
                        comments_parts.append(f"{q}: {a}")
                order_elements_text = "; ".join(comments_parts)
            
            # Extract only comment text from valueText (after "Оставьте свой комментарий" phrases)
            # This extracts the actual defect descriptions without form questions
            comments_only = ""
            if value_text:
                # Pattern: "Оставьте свой комментарий...:" followed by text until next question or ";"
                comment_pattern = re.compile(
                    r'Оставьте свой комментарий[^:]*:\s*([^;]+)',
                    re.IGNORECASE
                )
                matches = comment_pattern.findall(value_text)
                if matches:
                    # Clean up each match and join
                    cleaned = [m.strip() for m in matches if m.strip()]
                    comments_only = "; ".join(cleaned)
            
            row = {
                "id": order.get("id"),
                "serviceId": order.get("serviceId"),
                "serviceInternalTitle": order.get("serviceInternalTitle") or "",
                "ФИО": order.get("customerFullName") or order.get("customerShortName") or "",
                "Телефон": order.get("customerPhoneNumber") or "",
                "address": address,
                "placeNumber": order.get("placeNumber") or "",
                "placeId": order.get("placeId") or "",
                "placeExtId": order.get("placeExtId") or "",
                "title": order.get("serviceTitle") or "",
                "valueString": " | ".join(value_strings) if value_strings else "",
                "valueText": value_text,
                "commentsOnly": comments_only,
                "orderElements": order_elements_text,
                "Фото": photos,
                "extId": order.get("extId"),
                "createdAt": created_at_str,
            }
            data.append(row)
        
        return self._write_to_excel_ordered(data, output_path, "Orders", 
            ["id", "serviceId", "serviceInternalTitle", "ФИО", "Телефон", "address", "placeNumber", "placeId", "placeExtId", "title", "valueString", "valueText", "commentsOnly", "orderElements", "Фото", "extId", "createdAt"])
    
    def _write_to_excel_ordered(
        self, 
        data: list[dict], 
        output_path: Path, 
        sheet_name: str,
        headers: list[str]
    ) -> Path:
        """Write list of dicts to Excel file with specific column order."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
        logger.info(f"Exported {len(data)} records to {output_path}")
        return output_path
    
    async def export_payments(
        self,
        output_path: Path,
        date_time: Optional[str] = None,
    ) -> Path:
        """Export payments to Excel."""
        data = await self.client.get_payments(date_time=date_time)
        return self._write_to_excel(data, output_path, "Payments")
    
    async def export_orders_raw(
        self,
        output_path: Path,
        building_id: Optional[int] = None,
        created_at: Optional[str] = None,
        limit: int = 200,  # Limit for debugging
    ) -> Path:
        """Export orders with ALL fields (raw data for debugging). Limited to first N records."""
        # Get just first page for debugging
        params = {"fromRow": 0}
        if building_id:
            params["buildingId"] = building_id
        if created_at:
            params["createdAt"] = created_at
        
        data = await self.client._request("GET", "orders", params=params)
        items = data.get("items", [])[:limit]
        
        return self._write_to_excel(items, output_path, "Orders_Raw")
    
    async def export_order_comments(
        self,
        output_path: Path,
        order_ids: list[int],
    ) -> Path:
        """Export comments for specific orders."""
        all_comments = []
        for order_id in order_ids[:50]:  # Limit to 50 orders
            comments = await self.client.get_order_comments(order_id)
            for comment in comments:
                comment["orderId"] = order_id
                all_comments.append(comment)
        return self._write_to_excel(all_comments, output_path, "Order_Comments")
    
    async def export_acceptance_results(
        self,
        output_path: Path,
        building_id: Optional[int] = None,
    ) -> Path:
        """Export acceptance results (приёмка помещений)."""
        data = await self.client.get_acceptance_results(building_id=building_id)
        return self._write_to_excel(data, output_path, "Acceptance_Results")
    
    async def export_acceptance_defects(
        self,
        output_path: Path,
        building_id: Optional[int] = None,
    ) -> Path:
        """Export acceptance defects."""
        data = await self.client.get_acceptance_defects(building_id=building_id)
        return self._write_to_excel(data, output_path, "Acceptance_Defects")
    
    async def export_columns_list(self, output_path: Path) -> Path:
        """Export available columns list for orders."""
        data = await self.client.get_orders_export_columns()
        if isinstance(data, list):
            return self._write_to_excel(data, output_path, "Export_Columns")
        elif isinstance(data, dict):
            return self._write_to_excel([data], output_path, "Export_Columns")
        return self._write_to_excel([], output_path, "Export_Columns")
