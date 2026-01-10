"""Excel file writer for outputting classified defects.

Implements Requirements 7.1, 7.2, 7.3:
- Creates new XLSX file with all classified defects
- Includes all original columns plus "Категория дефекта"
- Saves to results/{job_id}_processed.xlsx
"""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.schemas import ExpandedRow


class ExcelWriterError(Exception):
    """Base exception for ExcelWriter errors."""
    pass


class ExcelWriter:
    """Writes classified defect data to Excel files.
    
    Creates output files with all original columns plus the
    "Категория дефекта" column containing the assigned category.
    """
    
    CATEGORY_COLUMN_NAME = "Категория дефекта"
    COMMENT_COLUMN_NAME = "КОММЕНТАРИЙ"
    
    def write_result(
        self,
        rows: list[ExpandedRow],
        output_path: str | Path,
        original_headers: Optional[list[str]] = None,
    ) -> Path:
        """Write classified rows to an xlsx file.
        
        Args:
            rows: List of ExpandedRow objects with categories assigned
            output_path: Path where to save the output file
            original_headers: Optional list of headers to preserve order.
                            If not provided, headers are extracted from first row.
                            
        Returns:
            Path to the created file
            
        Raises:
            ExcelWriterError: If file cannot be written
        """
        output_path = Path(output_path)
        
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            
            if sheet is None:
                raise ExcelWriterError("Failed to create worksheet")
            
            # Determine headers
            headers = self._get_headers(rows, original_headers)
            
            # Write headers
            self._write_headers(sheet, headers)
            
            # Write data rows
            self._write_rows(sheet, rows, headers)
            
            # Save workbook
            workbook.save(output_path)
            workbook.close()
            
            return output_path
            
        except ExcelWriterError:
            raise
        except Exception as e:
            raise ExcelWriterError(f"Failed to write file: {e}")
    
    def _get_headers(
        self,
        rows: list[ExpandedRow],
        original_headers: Optional[list[str]] = None,
    ) -> list[str]:
        """Determine the column headers for output file.
        
        Headers include all original columns plus "Категория дефекта".
        
        Args:
            rows: List of ExpandedRow objects
            original_headers: Optional predefined header order
            
        Returns:
            List of header names including category column
        """
        if original_headers:
            headers = list(original_headers)
        elif rows:
            # Extract headers from first row's original_data
            headers = list(rows[0].original_data.keys())
        else:
            # Empty result - just return category column
            return [self.CATEGORY_COLUMN_NAME]
        
        # Add category column if not present
        if self.CATEGORY_COLUMN_NAME not in headers:
            headers.append(self.CATEGORY_COLUMN_NAME)
        
        return headers
    
    def _write_headers(self, sheet: Worksheet, headers: list[str]) -> None:
        """Write header row to worksheet.
        
        Args:
            sheet: openpyxl worksheet
            headers: List of column headers
        """
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
    
    def _write_rows(
        self,
        sheet: Worksheet,
        rows: list[ExpandedRow],
        headers: list[str],
    ) -> None:
        """Write data rows to worksheet.
        
        Args:
            sheet: openpyxl worksheet
            rows: List of ExpandedRow objects
            headers: List of column headers (determines column order)
        """
        for row_idx, expanded_row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                if header == self.CATEGORY_COLUMN_NAME:
                    # Write category from ExpandedRow
                    value = expanded_row.category or ""
                else:
                    # Write from original_data
                    value = expanded_row.original_data.get(header)
                
                sheet.cell(row=row_idx, column=col_idx, value=value)


def get_output_path(job_id: str, results_dir: str | Path) -> Path:
    """Generate the output file path for a job.
    
    Follows the convention: results/{job_id}_processed.xlsx
    
    Args:
        job_id: Unique job identifier
        results_dir: Path to results directory
        
    Returns:
        Path to output file
    """
    results_dir = Path(results_dir)
    return results_dir / f"{job_id}_processed.xlsx"
