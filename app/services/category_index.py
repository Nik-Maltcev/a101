"""Category index for fuzzy search of defect categories.

Implements Requirements 5.1, 5.2, 5.3, 6.1:
- Loads category reference from data/categories.xlsx
- Builds fuzzy index for fast search
- Finds top-N similar categories
- Tracks file changes and rebuilds index
"""

import os
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from rapidfuzz import fuzz, process


class CategoryIndexError(Exception):
    """Base exception for CategoryIndex errors."""
    pass


class CategoryFileNotFoundError(CategoryIndexError):
    """Raised when the categories file is not found."""
    pass


class CategoryIndex:
    """Index for fast fuzzy search of defect categories.
    
    Uses rapidfuzz for efficient fuzzy string matching.
    Automatically tracks file changes and rebuilds index when needed.
    """
    
    def __init__(self, categories_path: str | Path):
        """Initialize CategoryIndex.
        
        Args:
            categories_path: Path to the categories.xlsx file
        """
        self._categories_path = Path(categories_path)
        self._categories: list[str] = []
        self._file_mtime: Optional[float] = None
        self._index_built = False
    
    @property
    def categories(self) -> list[str]:
        """Get the list of loaded categories."""
        return self._categories.copy()
    
    @property
    def is_loaded(self) -> bool:
        """Check if categories are loaded and index is built."""
        return self._index_built and len(self._categories) > 0
    
    def load_categories(self) -> list[str]:
        """Load categories from the Excel file.
        
        Reads the "Наименование" column (3rd column) as category names.
        
        Returns:
            List of category strings
            
        Raises:
            CategoryFileNotFoundError: If file doesn't exist
            CategoryIndexError: If file cannot be read
        """
        import re
        
        if not self._categories_path.exists():
            raise CategoryFileNotFoundError(
                f"Categories file not found: {self._categories_path}"
            )
        
        try:
            workbook = load_workbook(
                filename=self._categories_path, 
                read_only=True, 
                data_only=True
            )
            sheet = workbook.active
            
            if sheet is None:
                raise CategoryIndexError("No active sheet found in categories file")

            categories = []
            
            # Read "Наименование" column (3rd column, index 2), skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column C (index 2) is "Наименование"
                if len(row) >= 3:
                    cell_value = row[2]  # Наименование column
                    if cell_value is not None:
                        category = str(cell_value).strip()
                        if category:
                            categories.append(category)
            
            workbook.close()
            
            self._categories = categories
            self._file_mtime = os.path.getmtime(self._categories_path)
            
            return categories
            
        except CategoryIndexError:
            raise
        except Exception as e:
            raise CategoryIndexError(f"Failed to load categories: {e}")
    
    def build_index(self) -> None:
        """Build the fuzzy search index.
        
        Loads categories if not already loaded, then prepares
        the index for fast fuzzy matching.
        
        Raises:
            CategoryIndexError: If categories cannot be loaded
        """
        if not self._categories:
            self.load_categories()
        
        # rapidfuzz doesn't require explicit index building,
        # but we mark the index as ready for use
        self._index_built = True
    
    def find_top_n(self, text: str, n: int = 10) -> list[str]:
        """Find the N most similar categories for the given text.
        
        Uses multiple fuzzy string matching strategies to find categories 
        that best match the input text.
        
        Args:
            text: The defect text to match against categories
            n: Number of top matches to return (default: 10)
            
        Returns:
            List of up to N category strings, ordered by similarity
            
        Raises:
            CategoryIndexError: If index is not built
        """
        import logging
        logger = logging.getLogger(__name__)
        
        if not self._index_built:
            raise CategoryIndexError(
                "Index not built. Call build_index() first."
            )
        
        if not self._categories:
            return []
        
        if not text or not text.strip():
            # Return first N categories for empty input
            return self._categories[:n]
        
        query = text.strip().lower()
        
        # Extract key words from query for better matching
        # Common construction terms to look for
        key_terms = self._extract_key_terms(query)
        
        # Strategy 1: token_set_ratio - good for word overlap regardless of order
        results_token_set = process.extract(
            query=query,
            choices=self._categories,
            scorer=fuzz.token_set_ratio,
            limit=n * 2
        )
        
        # Strategy 2: partial_ratio - good for substring matches
        results_partial = process.extract(
            query=query,
            choices=self._categories,
            scorer=fuzz.partial_ratio,
            limit=n
        )
        
        # Strategy 3: WRatio - weighted ratio, good general purpose
        results_wratio = process.extract(
            query=query,
            choices=self._categories,
            scorer=fuzz.WRatio,
            limit=n
        )
        
        # Strategy 4: Search by key terms if found
        term_matches = []
        if key_terms:
            for term in key_terms:
                term_results = process.extract(
                    query=term,
                    choices=self._categories,
                    scorer=fuzz.partial_ratio,
                    limit=n // 2
                )
                term_matches.extend(term_results)
        
        # Combine all results with weighted scoring
        score_map = {}
        
        # Weight: token_set_ratio results (best for construction defects)
        for match, score, idx in results_token_set:
            if match not in score_map:
                score_map[match] = 0
            score_map[match] += score * 1.2  # Higher weight
        
        # Weight: partial_ratio results
        for match, score, idx in results_partial:
            if match not in score_map:
                score_map[match] = 0
            score_map[match] += score * 1.0
        
        # Weight: WRatio results
        for match, score, idx in results_wratio:
            if match not in score_map:
                score_map[match] = 0
            score_map[match] += score * 0.8
        
        # Weight: term matches (bonus for key term matches)
        for match, score, idx in term_matches:
            if match not in score_map:
                score_map[match] = 0
            score_map[match] += score * 0.5  # Bonus
        
        # Sort by combined score
        sorted_matches = sorted(score_map.items(), key=lambda x: x[1], reverse=True)
        
        # Take top N
        result = [match for match, score in sorted_matches[:n]]
        
        # Log for debugging (first few defects only)
        if len(query) < 100:
            logger.debug(f"Fuzzy search for '{query[:50]}...': top 5 = {result[:5]}")
        
        return result
    
    def _extract_key_terms(self, text: str) -> list[str]:
        """Extract key construction terms from text for better matching."""
        # Common construction terms in Russian
        key_patterns = [
            'окно', 'окон', 'оконн', 'стеклопакет', 'рама', 'пвх', 'створк',
            'дверь', 'двер', 'входн',
            'стен', 'стена', 'кладк', 'штукатурк',
            'пол', 'полов', 'ламинат', 'плитк', 'паркет',
            'потолок', 'потолоч',
            'электр', 'розетк', 'выключател', 'провод', 'кабел',
            'водоснабж', 'канализац', 'сантехник', 'труб', 'кран', 'смесител',
            'отоплен', 'радиатор', 'батаре',
            'вентиляц', 'кондиционер',
            'балкон', 'лоджи',
            'царапин', 'трещин', 'скол', 'повреждени', 'дефект',
            'загрязнен', 'пятн',
            'протечк', 'влаг',
            'зазор', 'щел', 'неплотн',
            'откос', 'подоконник', 'отлив',
            'уплотнител', 'резинк', 'герметик',
            'фурнитур', 'ручк', 'петл', 'замок',
            'монтаж', 'установк',
            'обои', 'покраск', 'краск',
        ]
        
        found_terms = []
        text_lower = text.lower()
        
        for pattern in key_patterns:
            if pattern in text_lower:
                found_terms.append(pattern)
        
        return found_terms[:5]  # Limit to 5 most relevant terms
    
    def check_and_rebuild(self) -> bool:
        """Check if the categories file has changed and rebuild if needed.
        
        Compares the file modification time with the cached value.
        If the file has been modified, reloads categories and rebuilds index.
        
        Returns:
            True if index was rebuilt, False otherwise
            
        Raises:
            CategoryIndexError: If rebuild fails
        """
        if not self._categories_path.exists():
            raise CategoryFileNotFoundError(
                f"Categories file not found: {self._categories_path}"
            )
        
        current_mtime = os.path.getmtime(self._categories_path)
        
        # Rebuild if file was modified or index was never built
        if self._file_mtime is None or current_mtime > self._file_mtime:
            self.load_categories()
            self.build_index()
            return True
        
        return False
