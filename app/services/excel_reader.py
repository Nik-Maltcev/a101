"""Excel file reader for processing defect comments."""

from pathlib import Path
from typing import Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReaderError(Exception):
    """Base exception for ExcelReader errors."""
    pass


class CommentColumnNotFoundError(ExcelReaderError):
    """Raised when the comment column is not found."""
    pass


class ExcelReader:
    """Reads and parses Excel files with defect comments.
    
    Implements Requirements 2.1, 2.2, 2.3:
    - Opens XLSX files using openpyxl
    - Finds the "valueString" and "valueText" columns
    - Extracts all data rows
    """
    
    def read_file(self, file_path: str | Path) -> list[dict]:
        """Read xlsx file and return list of rows as dictionaries.
        
        Args:
            file_path: Path to the xlsx file
            
        Returns:
            List of dictionaries where keys are column headers
            
        Raises:
            ExcelReaderError: If file cannot be read
            CommentColumnNotFoundError: If КОММЕНТАРИЙ column not found
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ExcelReaderError(f"File not found: {file_path}")
        
        if file_path.suffix.lower() != ".xlsx":
            raise ExcelReaderError(f"Invalid file format: {file_path.suffix}")
        
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            if sheet is None:
                raise ExcelReaderError("No active sheet found in workbook")
            
            # Get headers from first row
            headers = self._get_headers(sheet)
            
            # Verify at least one of valueString or valueText exists
            has_value_string = any(h.upper() == "VALUESTRING" for h in headers)
            has_value_text = any(h.upper() == "VALUETEXT" for h in headers)
            
            if not has_value_string and not has_value_text:
                raise CommentColumnNotFoundError(
                    "Neither 'valueString' nor 'valueText' column found in file"
                )
            
            # Extract data rows
            rows = self._extract_rows(sheet, headers)
            
            workbook.close()
            return rows
            
        except CommentColumnNotFoundError:
            raise
        except ExcelReaderError:
            raise
        except Exception as e:
            raise ExcelReaderError(f"Failed to read file: {e}")
    
    def find_comment_column(self, sheet: Worksheet) -> Optional[int]:
        """Find the index of the comment column.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            Column index (0-based) or None if not found
        """
        # This method is deprecated but kept for backwards compatibility
        return None
    
    def _get_headers(self, sheet: Worksheet) -> list[str]:
        """Extract column headers from the first row.
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            List of header names
        """
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        
        if first_row is None:
            return []
        
        headers = []
        for cell_value in first_row:
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append("")
        
        return headers
    
    def _extract_rows(self, sheet: Worksheet, headers: list[str]) -> list[dict]:
        """Extract all data rows from the sheet.
        
        Args:
            sheet: openpyxl worksheet
            headers: List of column headers
            
        Returns:
            List of dictionaries with column headers as keys
        """
        rows = []
        
        # Skip header row, iterate from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            
            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    value = row[idx]
                    # Convert to string if not None, preserve None for empty cells
                    row_dict[header] = str(value) if value is not None else None
                else:
                    row_dict[header] = None
            
            rows.append(row_dict)
        
        return rows
